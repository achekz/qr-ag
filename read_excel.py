#!/usr/bin/env python3
"""
Script to read .xlsx files using different libraries
"""

import pandas as pd
import openpyxl
import os
import sys

def read_with_pandas(file_path):
    """Read Excel file using pandas"""
    try:
        # Read all sheets
        excel_file = pd.ExcelFile(file_path)
        print(f"Sheet names: {excel_file.sheet_names}")
        
        # Read first sheet
        df = pd.read_excel(file_path)
        print(f"\nFirst sheet data shape: {df.shape}")
        print(f"Columns: {list(df.columns)}")
        print("\nFirst 5 rows:")
        print(df.head())
        
        return df
    except Exception as e:
        print(f"Error reading with pandas: {e}")
        return None

def read_with_openpyxl(file_path):
    """Read Excel file using openpyxl"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        print(f"\nSheet names: {workbook.sheetnames}")
        
        # Read first sheet
        sheet = workbook.active
        print(f"Active sheet: {sheet.title}")
        print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")
        
        # Read first few rows
        print("\nFirst 5 rows of data:")
        for row in range(1, min(6, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(6, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
            print(f"Row {row}: {row_data}")
        
        return workbook
    except Exception as e:
        print(f"Error reading with openpyxl: {e}")
        return None

def main():
    if len(sys.argv) != 2:
        print("Usage: python read_excel.py <path_to_excel_file>")
        print("Example: python read_excel.py data.xlsx")
        return
    
    file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return
    
    if not file_path.lower().endswith(('.xlsx', '.xls')):
        print("Please provide an Excel file (.xlsx or .xls)")
        return
    
    print(f"Reading Excel file: {file_path}")
    print("=" * 50)
    
    # Try pandas first
    print("Using pandas:")
    df = read_with_pandas(file_path)
    
    print("\n" + "=" * 50)
    
    # Try openpyxl
    print("Using openpyxl:")
    workbook = read_with_openpyxl(file_path)

if __name__ == "__main__":
    main()
